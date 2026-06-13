from datetime import datetime
from io import BytesIO

from django.http import FileResponse
from rest_framework import status
from rest_framework.response import Response
from rest_framework.views import APIView

from apps.projects.serializers import ProyectoGradoSerializer
from apps.projects.views import proyectos_para
from apps.users.permissions import IsDocenteLike

GUINDO = '6B1D2F'

COLUMNAS = [
    ('Código', 'codigo', 18),
    ('Proyecto', 'titulo', 45),
    ('Estudiante', 'estudiante_nombre', 28),
    ('Tutor', 'tutor_nombre', 28),
    ('Estado', 'estado_revision', 16),
    ('Etapa', 'etapa', 16),
    ('Obs. pendientes', 'observaciones_pendientes', 16),
]


def _filas_proyectos(request):
    qs = proyectos_para(request.user).order_by('-updated_at')
    data = ProyectoGradoSerializer(qs, many=True).data
    filas = []
    for proyecto in data:
        filas.append([
            proyecto['codigo'],
            proyecto['titulo'],
            proyecto['estudiante_nombre'],
            proyecto['tutor_nombre'] or '—',
            proyecto['estado_revision'],
            proyecto['etapa'],
            proyecto['observaciones_pendientes'],
        ])
    return filas


def _xlsx_response(filas):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    wb = Workbook()
    ws = wb.active
    ws.title = 'Proyectos'

    header_fill = PatternFill('solid', fgColor=GUINDO)
    header_font = Font(color='FFFFFF', bold=True)
    for col, (titulo, _, ancho) in enumerate(COLUMNAS, start=1):
        celda = ws.cell(row=1, column=col, value=titulo)
        celda.fill = header_fill
        celda.font = header_font
        ws.column_dimensions[celda.column_letter].width = ancho

    for fila in filas:
        ws.append(fila)

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    nombre = f"proyectos_{datetime.now():%Y%m%d_%H%M}.xlsx"
    return FileResponse(
        buffer,
        as_attachment=True,
        filename=nombre,
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )


def _pdf_response(filas):
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.lib.units import cm
    from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

    buffer = BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=landscape(A4),
        leftMargin=1.2 * cm,
        rightMargin=1.2 * cm,
        topMargin=1.2 * cm,
        bottomMargin=1.2 * cm,
    )
    styles = getSampleStyleSheet()
    titulo = Paragraph(
        '<font color="#6B1D2F"><b>AcademicFlow — Gestión de Proyectos</b></font>',
        styles['Title'],
    )
    fecha = Paragraph(
        f"Generado el {datetime.now():%d/%m/%Y %H:%M}", styles['Normal']
    )

    encabezados = [titulo_col for titulo_col, _, _ in COLUMNAS]
    body_style = styles['BodyText']
    body_style.fontSize = 8
    datos = [encabezados] + [
        [Paragraph(str(valor), body_style) for valor in fila] for fila in filas
    ]
    tabla = Table(datos, repeatRows=1)
    tabla.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#6B1D2F')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTSIZE', (0, 0), (-1, 0), 9),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1),
         [colors.white, colors.HexColor('#F3F3F3')]),
        ('GRID', (0, 0), (-1, -1), 0.4, colors.HexColor('#DAC0C2')),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ('TOPPADDING', (0, 0), (-1, -1), 4),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
    ]))

    doc.build([titulo, fecha, Spacer(1, 12), tabla])
    buffer.seek(0)
    nombre = f"proyectos_{datetime.now():%Y%m%d_%H%M}.pdf"
    return FileResponse(
        buffer, as_attachment=True, filename=nombre,
        content_type='application/pdf',
    )


class ProjectsExportView(APIView):
    permission_classes = [IsDocenteLike]

    def get(self, request):
        # 'format' está reservado por DRF (sufijo de renderer); se usa 'formato'
        formato = request.query_params.get('formato', 'xlsx')
        if formato not in ('xlsx', 'pdf'):
            return Response(
                {'detail': "El formato debe ser 'xlsx' o 'pdf'."},
                status=status.HTTP_400_BAD_REQUEST,
            )
        filas = _filas_proyectos(request)
        if formato == 'xlsx':
            return _xlsx_response(filas)
        return _pdf_response(filas)


class ReportExportView(APIView):
    """Exportación parametrizada del módulo de reportes (modal de exportación)."""

    permission_classes = [IsDocenteLike]

    def post(self, request):
        formato = request.data.get('formato', 'pdf')
        if formato not in ('xlsx', 'pdf'):
            return Response(
                {'detail': "El formato debe ser 'xlsx' o 'pdf'."},
                status=status.HTTP_400_BAD_REQUEST,
            )
        filas = _filas_proyectos(request)
        if formato == 'xlsx':
            return _xlsx_response(filas)
        return _pdf_response(filas)
